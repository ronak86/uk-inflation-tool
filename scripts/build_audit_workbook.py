import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname


ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "web" / "data" / "inflation.json"
OUTPUT_PATH = ROOT / "outputs" / "uk_inflation_calculation_audit.xlsx"


BLUE = "1F5F9F"
LIGHT_BLUE = "DDEBF7"
PALE_BLUE = "EAF4FB"
WHITE = "FFFFFF"
GREY = "F2F2F2"


def load_series():
    data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    return data["series"]


def col_letter(col):
    return get_column_letter(col)


def sheet_ref(sheet, cell):
    return f"{quote_sheetname(sheet)}!{cell}"


def month_parts(month):
    year, month_num = month.split("-")
    return int(year), int(month_num)


def month_label(month):
    year, month_num = month_parts(month)
    names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{names[month_num - 1]} {str(year)[-2:]}"


def safe_title(text):
    return text[:31]


def find_month(months, year, month_num):
    key = f"{year:04d}-{month_num:02d}"
    try:
        return months.index(key)
    except ValueError:
        return -1


def leaf_level(items):
    return max(item["level"] for item in items)


def descendants_at_leaf(items, item_id, leaf):
    out = []
    stack = list(items[item_id].get("children", []))
    while stack:
        child_id = stack.pop(0)
        child = items[child_id]
        if child["level"] == leaf:
            out.append(child_id)
        stack = list(child.get("children", [])) + stack
    return out


def apply_header_style(ws, row=1):
    fill = PatternFill("solid", fgColor=BLUE)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def freeze_and_filter(ws, max_row, max_col):
    ws.freeze_panes = "H2"
    ws.auto_filter.ref = f"A1:{col_letter(max_col)}{max_row}"


class ModelBuilder:
    def __init__(self, wb, family, payload):
        self.wb = wb
        self.family = family
        self.payload = payload
        self.items = payload["items"]
        self.months = payload["months"]
        self.n_months = len(self.months)
        self.leaf = leaf_level(self.items)
        self.input_sheet = safe_title(f"{family} Inputs")
        self.input_row = {}
        self.sheet_rows = {}

    def build(self):
        self.build_inputs()
        self.build_weights()
        self.build_price_change("mom")
        self.build_price_change("yoy")
        self.build_contribution("mom")
        self.build_contribution("yoy")
        self.build_checks()

    def input_price_cell(self, item_id, month_index):
        row = self.input_row[item_id]
        col = 8 + month_index
        return sheet_ref(self.input_sheet, f"{col_letter(col)}{row}")

    def input_weight_cell(self, item_id, month_index):
        row = self.input_row[item_id]
        col = 8 + self.n_months + 1 + month_index
        return sheet_ref(self.input_sheet, f"{col_letter(col)}{row}")

    def overall_3dp_cell_or_price(self, month_index):
        overall = self.payload.get("overall3dp")
        if not overall:
            return self.input_price_cell(0, month_index)
        month = self.months[month_index]
        if month not in overall.get("months", []):
            return self.input_price_cell(0, month_index)
        overall_month_index = overall["months"].index(month)
        col = 8 + self.n_months + 1 + self.n_months + 1 + overall_month_index
        return sheet_ref(self.input_sheet, f"{col_letter(col)}{self.input_row[0]}")

    def unchained_index(self, item_id, month_index):
        year, month_num = month_parts(self.months[month_index])
        current = self.input_price_cell(item_id, month_index)
        if month_num == 1:
            base_index = find_month(self.months, year - 1, 1 if self.family == "RPI" else 12)
            if base_index < 0:
                return "NA()"
            return f"({current}/{self.input_price_cell(item_id, base_index)})*100"
        january = find_month(self.months, year, 1)
        return f"({current}/{self.input_price_cell(item_id, january)})*100"

    def unchained_current_jan(self, item_id, month_index):
        year, _ = month_parts(self.months[month_index])
        january = find_month(self.months, year, 1)
        return f"({self.input_price_cell(item_id, month_index)}/{self.input_price_cell(item_id, january)})*100"

    def monthly_weight_index(self, month_index):
        _, month_num = month_parts(self.months[month_index])
        return month_index - 1 if self.family == "RPI" and month_num == 1 else month_index

    def build_inputs(self):
        ws = self.wb.create_sheet(self.input_sheet)
        headers = ["ID", "Name", "Level", "Parent ID", "Weight Code", "Price Code", "Prefix"]
        headers += [f"Price {month_label(m)}" for m in self.months]
        headers += [""] + [f"Weight {month_label(m)}" for m in self.months]
        overall = self.payload.get("overall3dp")
        if overall:
            headers += [""] + [f"Overall 3dp {month_label(m)}" for m in overall["months"]]
        ws.append(headers)

        for item_id, item in enumerate(self.items):
            self.input_row[item_id] = item_id + 2
            row = [
                item_id,
                item["name"],
                item["level"],
                item.get("parentId"),
                item.get("weightCode", ""),
                item.get("priceCode", ""),
                item["name"].strip().split(" ")[0] if item["level"] > 0 else "",
            ]
            row += item["prices"]
            row += [""] + item["weights"]
            if overall:
                row += [""] + (overall["prices"] if item_id == 0 else [""] * len(overall["months"]))
            ws.append(row)

        apply_header_style(ws)
        freeze_and_filter(ws, len(self.items) + 1, len(headers))
        ws.column_dimensions["B"].width = 44
        for col in range(1, len(headers) + 1):
            if col != 2:
                ws.column_dimensions[col_letter(col)].width = 12
        ws.sheet_view.showGridLines = False

    def build_meta_headers(self, ws, title):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13, color=BLUE)
        headers = ["ID", "Name", "Level", "Parent ID", "Weight Code", "Price Code", "Formula Scope"]
        headers += [month_label(m) for m in self.months]
        ws.append(headers)
        apply_header_style(ws, 2)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[col_letter(col)].width = 13
        ws.column_dimensions["B"].width = 44
        ws.freeze_panes = "H3"

    def write_meta_row(self, ws, item_id, row_num, scope):
        item = self.items[item_id]
        ws.cell(row_num, 1, item_id)
        ws.cell(row_num, 2, item["name"])
        ws.cell(row_num, 3, item["level"])
        ws.cell(row_num, 4, item.get("parentId"))
        ws.cell(row_num, 5, item.get("weightCode", ""))
        ws.cell(row_num, 6, item.get("priceCode", ""))
        ws.cell(row_num, 7, scope)

    def formula_sheet_setup(self, ws, title):
        self.build_meta_headers(ws, title)
        sheet_rows = {}
        for item_id, item in enumerate(self.items):
            row_num = item_id + 3
            sheet_rows[item_id] = row_num
            scope = "Leaf formula" if item["level"] == self.leaf else "Roll-up sum of leaf rows"
            self.write_meta_row(ws, item_id, row_num, scope)
        self.sheet_rows[ws.title] = sheet_rows
        return sheet_rows

    def leaf_row_refs(self, sheet_rows, item_id, month_col):
        if self.items[item_id]["level"] == self.leaf:
            leaf_ids = [item_id]
        elif item_id == 0:
            leaf_ids = [i for i, item in enumerate(self.items) if item["level"] == self.leaf]
        else:
            leaf_ids = descendants_at_leaf(self.items, item_id, self.leaf)
        return [f"{col_letter(month_col)}{sheet_rows[leaf_id]}" for leaf_id in leaf_ids]

    def build_weights(self):
        title = safe_title(f"{self.family} Weights %")
        ws = self.wb.create_sheet(title)
        self.build_meta_headers(ws, f"{self.family} weights shown as percent of full basket")
        for item_id, item in enumerate(self.items):
            row_num = item_id + 3
            self.write_meta_row(ws, item_id, row_num, "Weight / 10")
            for month_index in range(self.n_months):
                col = 8 + month_index
                ws.cell(row_num, col, f"={self.input_weight_cell(item_id, month_index)}/10")
        self.finish_calc_sheet(ws)

    def build_price_change(self, horizon):
        label = "MoM" if horizon == "mom" else "YoY"
        title = safe_title(f"{self.family} Price {label}")
        ws = self.wb.create_sheet(title)
        self.build_meta_headers(ws, f"{self.family} {label} price change, %")
        lag = 1 if horizon == "mom" else 12
        for item_id, item in enumerate(self.items):
            row_num = item_id + 3
            self.write_meta_row(ws, item_id, row_num, f"=current / prior {label} - 1")
            for month_index in range(self.n_months):
                col = 8 + month_index
                if month_index < lag:
                    ws.cell(row_num, col, "")
                    continue
                current = self.overall_3dp_cell_or_price(month_index) if item_id == 0 else self.input_price_cell(item_id, month_index)
                previous = self.overall_3dp_cell_or_price(month_index - lag) if item_id == 0 else self.input_price_cell(item_id, month_index - lag)
                ws.cell(row_num, col, f"=IFERROR(({current}/{previous}-1)*100,\"\")")
        self.finish_calc_sheet(ws)

    def monthly_leaf_formula(self, item_id, month_index):
        if month_index <= 0:
            return ""
        _, month_num = month_parts(self.months[month_index])
        weight_idx = self.monthly_weight_index(month_index)
        if month_num == 1:
            item_current = self.unchained_index(item_id, month_index)
            if self.family == "RPI":
                item_previous = self.unchained_current_jan(item_id, month_index - 1)
                all_previous = self.unchained_current_jan(0, month_index - 1)
            else:
                item_previous = "100"
                all_previous = "100"
        else:
            item_current = self.unchained_current_jan(item_id, month_index)
            item_previous = self.unchained_current_jan(item_id, month_index - 1)
            all_previous = self.unchained_current_jan(0, month_index - 1)
        weight = self.input_weight_cell(item_id, weight_idx)
        return f"=IFERROR((({item_current})/({item_previous})-1)*100*(({item_previous})/({all_previous}))*({weight}/1000),\"\")"

    def annual_leaf_formula(self, item_id, month_index):
        if month_index < 12:
            return ""
        year, month_num = month_parts(self.months[month_index])
        previous_month = find_month(self.months, year - 1, month_num)
        if previous_month < 0:
            return ""
        if self.family == "RPI":
            previous_jan = find_month(self.months, year - 1, 1)
            current_jan = find_month(self.months, year, 1)
            if previous_jan < 0 or current_jan < 0:
                return ""
            denom = self.unchained_current_jan(0, previous_month)
            all_current_jan = self.unchained_index(0, current_jan)
            item_prev_month = self.unchained_current_jan(item_id, previous_month)
            item_current_jan = self.unchained_index(item_id, current_jan)
            item_current_month = "100" if month_num == 1 else self.unchained_current_jan(item_id, month_index)
            previous_weight = f"({self.input_weight_cell(item_id, previous_jan)}/1000)"
            current_weight = f"({self.input_weight_cell(item_id, current_jan)}/1000)"
            term_one = f"{previous_weight}*((({item_current_jan})-({item_prev_month}))/({denom}))*100"
            term_two = "0" if month_num == 1 else f"{current_weight}*((({item_current_month})-100)/({denom}))*(({all_current_jan})/100)*100"
            return f"=IFERROR(({term_one})+({term_two}),\"\")"

        previous_dec = find_month(self.months, year - 1, 12)
        current_jan = find_month(self.months, year, 1)
        current_feb = find_month(self.months, year, 2)
        previous_feb = find_month(self.months, year - 1, 2)
        if previous_dec < 0 or current_jan < 0 or previous_feb < 0:
            return ""
        denom = self.unchained_current_jan(0, previous_month)
        all_previous_dec = self.unchained_current_jan(0, previous_dec)
        all_current_jan = self.unchained_index(0, current_jan)
        item_previous_month = self.unchained_current_jan(item_id, previous_month)
        item_previous_dec = self.unchained_current_jan(item_id, previous_dec)
        item_current_jan = self.unchained_index(item_id, current_jan)
        item_current_month = self.unchained_current_jan(item_id, month_index)
        previous_weight = f"({self.input_weight_cell(item_id, previous_feb)}/1000)"
        january_weight = f"({self.input_weight_cell(item_id, current_jan)}/1000)"
        term_one = f"{previous_weight}*((({item_previous_dec})-({item_previous_month}))/({denom}))*100"
        term_two = f"{january_weight}*((({item_current_jan})-100)/({denom}))*({all_previous_dec})"
        if month_num == 1 or current_feb < 0:
            term_three = "0"
        else:
            term_three = (
                f"({self.input_weight_cell(item_id, current_feb)}/1000)"
                f"*(({item_current_month})-100)/({denom})"
                f"*(({all_current_jan})/100)*({all_previous_dec})"
            )
        return f"=IFERROR(({term_one})+({term_two})+({term_three}),\"\")"

    def build_contribution(self, horizon):
        label = "MoM" if horizon == "mom" else "YoY"
        title = safe_title(f"{self.family} {label} Ctrb")
        ws = self.wb.create_sheet(title)
        sheet_rows = self.formula_sheet_setup(ws, f"{self.family} {label} contribution, basis points")
        for item_id, item in enumerate(self.items):
            row_num = sheet_rows[item_id]
            for month_index in range(self.n_months):
                col = 8 + month_index
                if item["level"] == self.leaf:
                    formula = self.monthly_leaf_formula(item_id, month_index) if horizon == "mom" else self.annual_leaf_formula(item_id, month_index)
                else:
                    refs = self.leaf_row_refs(sheet_rows, item_id, col)
                    formula = f"=SUM({','.join(refs)})" if refs else ""
                ws.cell(row_num, col, formula)
        self.finish_calc_sheet(ws)

    def build_checks(self):
        ws = self.wb.create_sheet(safe_title(f"{self.family} Checks"))
        headers = ["Metric"] + [month_label(m) for m in self.months]
        ws.append(headers)
        apply_header_style(ws)
        rows = [
            ("Headline MoM, %", "mom_headline"),
            ("Sum MoM contrib, bp", "mom_sum"),
            ("MoM error, bp", "mom_error"),
            ("Headline YoY, %", "yoy_headline"),
            ("Sum YoY contrib, bp", "yoy_sum"),
            ("YoY error, bp", "yoy_error"),
        ]
        for row_index, (label, key) in enumerate(rows, start=2):
            ws.cell(row_index, 1, label)
            for month_index in range(self.n_months):
                col = 2 + month_index
                if key == "mom_headline":
                    if month_index == 0:
                        formula = ""
                    else:
                        formula = f"=IFERROR(({self.overall_3dp_cell_or_price(month_index)}/{self.overall_3dp_cell_or_price(month_index - 1)}-1)*100,\"\")"
                elif key == "yoy_headline":
                    if month_index < 12:
                        formula = ""
                    else:
                        formula = f"=IFERROR(({self.overall_3dp_cell_or_price(month_index)}/{self.overall_3dp_cell_or_price(month_index - 12)}-1)*100,\"\")"
                elif key == "mom_sum":
                    formula = f"='{self.family} MoM Ctrb'!{col_letter(8 + month_index)}3"
                elif key == "yoy_sum":
                    formula = f"='{self.family} YoY Ctrb'!{col_letter(8 + month_index)}3"
                elif key == "mom_error":
                    formula = f"=IFERROR({col_letter(col)}3-{col_letter(col)}2,\"\")"
                else:
                    formula = f"=IFERROR({col_letter(col)}6-{col_letter(col)}5,\"\")"
                ws.cell(row_index, col, formula)
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 24
        for col in range(2, self.n_months + 2):
            ws.column_dimensions[col_letter(col)].width = 10
        ws.sheet_view.showGridLines = False

    def finish_calc_sheet(self, ws):
        max_row = len(self.items) + 2
        max_col = 7 + self.n_months
        freeze_and_filter(ws, max_row, max_col)
        for row in range(3, max_row + 1):
            level = ws.cell(row, 3).value
            if level == 0:
                fill = PatternFill("solid", fgColor=BLUE)
                font = Font(color=WHITE, bold=True)
            elif level == 1:
                fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                font = Font(color="000000")
            elif row % 2 == 0:
                fill = PatternFill("solid", fgColor=PALE_BLUE)
                font = Font(color="000000")
            else:
                fill = PatternFill("solid", fgColor=WHITE)
                font = Font(color="000000")
            for col in range(1, max_col + 1):
                ws.cell(row, col).fill = fill
                ws.cell(row, col).font = font
        for col in range(8, max_col + 1):
            for row in range(3, max_row + 1):
                ws.cell(row, col).number_format = "0.000000"
        ws.sheet_view.showGridLines = False


def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    lines = [
        ("UK Inflation Calculation Audit Workbook", "Title"),
        ("Purpose", "Header"),
        ("This workbook exposes the core CPI, CPIH and RPI calculations as Excel formulas so individual cells can be inspected, traced and edited.", "Text"),
        ("Structure", "Header"),
        ("Inputs tabs contain the rounded public index and weight data used by the app. Weight columns are in ONS weight units, so divide by 10 to show percent.", "Text"),
        ("Price tabs calculate month-on-month and year-on-year price changes. CPI/CPIH headline rows use the 3dp overall index where available; RPI uses the headline index available in the source data.", "Text"),
        ("Contribution tabs calculate contributions at the most detailed available level first, then parent rows sum those detailed contribution rows.", "Text"),
        ("Checks tabs compare headline price change with the summed detailed contributions. Errors are shown in basis points.", "Text"),
        ("Formula Notes", "Header"),
        ("CPI/CPIH leaves use the ONS chain-linked contribution method with February/current-year weights where applicable. RPI leaves use the RPI January chain-linking treatment, with January monthly contributions using the prior month weight.", "Text"),
        ("Custom sector baskets in the web app dynamically filter leaves and renormalise weights. This workbook focuses on the full CPI, CPIH and RPI baskets as a transparent audit trail.", "Text"),
    ]
    for row, (text, kind) in enumerate(lines, start=1):
        cell = ws.cell(row, 1, text)
        if kind == "Title":
            cell.font = Font(bold=True, size=16, color=BLUE)
        elif kind == "Header":
            cell.font = Font(bold=True, size=12, color=BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws.sheet_view.showGridLines = False


def main():
    series = load_series()
    wb = Workbook()
    build_readme(wb)
    for family in ["CPI", "CPIH", "RPI"]:
        if family in series:
            ModelBuilder(wb, family, series[family]).build()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
