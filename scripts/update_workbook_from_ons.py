from __future__ import annotations

import argparse
from copy import copy
import re
import shutil
import sys
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "Weights And Prices.xlsx"
DEFAULT_DOWNLOAD_DIR = ROOT / "raw_ons"
ONS_URL = (
    "https://www.ons.gov.uk/file?uri=/economy/inflationandpriceindices/"
    "datasets/consumerpriceinflation/current/consumerpriceinflationdetailedreferencetables.xlsx"
)

SERIES_CONFIG = {
    "CPIH": {
        "price_table": "Table 7",
        "weight_table": "Table 9",
        "price_start_row": 7,
        "weight_header_row": 7,
        "weight_start_row": 8,
    },
    "CPI": {
        "price_table": "Table 16",
        "weight_table": "Table 18",
        "price_start_row": 7,
        "weight_header_row": 6,
        "weight_start_row": 7,
    },
    "RPI": {
        "price_table": "Table 27",
        "weight_table": "Table 29",
        "price_start_row": 9,
        "weight_header_row": 6,
        "weight_start_row": 7,
    },
}

OVERALL_3DP = {
    "D7BT": {"table": "Table 38"},
    "L522": {"table": "Table 37"},
}


def clean(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def clean_code(value: Any) -> str | None:
    value = clean(value)
    if value in (None, "", "#N/A"):
        return None
    return str(value).strip()


def month_from_parts(year: Any, month_name: Any) -> datetime | None:
    if not isinstance(year, int) or not isinstance(month_name, str):
        return None
    try:
        month_number = datetime.strptime(month_name.strip()[:3], "%b").month
    except ValueError:
        return None
    return datetime(year, month_number, 1)


def month_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    raise ValueError(f"Expected month header date, got {value!r}")


def parse_weight_header(value: Any) -> tuple[int, str]:
    text = str(value).replace("\n", " ").strip()
    year_match = re.search(r"(20\d{2}|19\d{2})", text)
    if not year_match:
        raise ValueError(f"Could not parse weight year from header {value!r}")
    year = int(year_match.group(1))
    lower = text.lower()
    if "jan" in lower:
        period = "jan"
    elif "feb" in lower or "dec" in lower:
        period = "feb-dec"
    else:
        period = "all"
    return year, period


def select_weight_column(raw_sheet, target_month: datetime, header_row: int) -> int:
    candidates: list[tuple[int, str, int]] = []
    for col in range(1, raw_sheet.max_column + 1):
        header = raw_sheet.cell(header_row, col).value
        if header is None:
            continue
        try:
            year, period = parse_weight_header(header)
        except ValueError:
            continue
        if year == target_month.year:
            candidates.append((year, period, col))

    if not candidates:
        raise ValueError(f"No weight column found for {target_month.year} on {raw_sheet.title}")

    wanted_period = "jan" if target_month.month == 1 else "feb-dec"
    for _, period, col in candidates:
        if period == wanted_period:
            return col
    for _, period, col in candidates:
        if period == "all":
            return col
    if target_month.month != 1:
        for _, period, col in candidates:
            if period == "feb-dec":
                return col
    raise ValueError(f"No usable weight column found for {target_month:%Y-%m} on {raw_sheet.title}")


def read_latest_price_series(raw_wb, table_name: str, start_row: int) -> tuple[datetime, dict[str, float]]:
    sheet = raw_wb[table_name]
    latest_col = None
    latest_month = None
    for col in range(1, sheet.max_column + 1):
        parsed = month_from_parts(sheet.cell(5 if table_name != "Table 27" else 7, col).value, sheet.cell(6 if table_name != "Table 27" else 8, col).value)
        if parsed:
            latest_col = col
            latest_month = parsed
    if latest_col is None or latest_month is None:
        raise ValueError(f"Could not find latest month in {table_name}")

    values: dict[str, float] = {}
    for row in range(start_row, sheet.max_row + 1):
        code = clean_code(sheet.cell(row, 2).value)
        value = sheet.cell(row, latest_col).value
        if code and isinstance(value, (int, float)):
            values[code] = float(value)
    return latest_month, values


def read_weight_series(raw_wb, table_name: str, target_month: datetime, header_row: int, start_row: int) -> dict[str, float]:
    sheet = raw_wb[table_name]
    value_col = select_weight_column(sheet, target_month, header_row)
    values: dict[str, float] = {}
    for row in range(start_row, sheet.max_row + 1):
        code = clean_code(sheet.cell(row, 2).value)
        value = sheet.cell(row, value_col).value
        if code and isinstance(value, (int, float)):
            values[code] = float(value)
    return values


def read_3dp_overall(raw_wb, code: str) -> tuple[datetime, float]:
    sheet = raw_wb[OVERALL_3DP[code]["table"]]
    code_col = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number != 6:
            continue
        for index, value in enumerate(row):
            if clean_code(value) == code:
                code_col = index
                break
        break
    if code_col is None:
        raise ValueError(f"Could not find {code} in {sheet.title}")

    latest_month = None
    latest_value = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number < 8 or len(row) <= code_col:
            continue
        date_value = row[1] if len(row) > 1 else None
        value = row[code_col]
        if isinstance(date_value, datetime) and isinstance(value, (int, float)):
            latest_month = datetime(date_value.year, date_value.month, 1)
            latest_value = float(value)
    if latest_month is None or latest_value is None:
        raise ValueError(f"Could not read latest {code} value from {sheet.title}")
    return latest_month, latest_value


def copy_column_style(sheet, source_col: int, target_col: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
    for row in range(1, sheet.max_row + 1):
        source = sheet.cell(row, source_col)
        target = sheet.cell(row, target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def ensure_month_column(sheet, target_month: datetime) -> int:
    target_key = target_month.strftime("%Y-%m")
    for col in range(5, sheet.max_column + 1):
        value = sheet.cell(1, col).value
        if isinstance(value, datetime) and value.strftime("%Y-%m") == target_key:
            return col

    last_month = sheet.cell(1, sheet.max_column).value
    if not isinstance(last_month, datetime):
        raise ValueError(f"Last header in {sheet.title} is not a date")
    if target_month <= last_month:
        raise ValueError(f"{target_month:%Y-%m} is not present and is not after existing last month {last_month:%Y-%m}")

    new_col = sheet.max_column + 1
    copy_column_style(sheet, sheet.max_column, new_col)
    sheet.cell(1, new_col).value = target_month
    return new_col


def update_series(workbook, raw_wb, series: str) -> dict[str, Any]:
    config = SERIES_CONFIG[series]
    latest_month, price_values = read_latest_price_series(raw_wb, config["price_table"], config["price_start_row"])
    weight_values = read_weight_series(raw_wb, config["weight_table"], latest_month, config["weight_header_row"], config["weight_start_row"])

    results = {
        "series": series,
        "month": latest_month.strftime("%Y-%m"),
        "prices_updated": 0,
        "weights_updated": 0,
        "missing_price_codes": [],
        "missing_weight_codes": [],
    }

    for kind, values, counter_name, missing_name in (
        ("Prices", price_values, "prices_updated", "missing_price_codes"),
        ("Weights", weight_values, "weights_updated", "missing_weight_codes"),
    ):
        sheet = workbook[f"{series}_{kind}"]
        month_col = ensure_month_column(sheet, latest_month)
        code_col = 4 if kind == "Prices" else 3
        for row in range(2, sheet.max_row + 1):
            code = clean_code(sheet.cell(row, code_col).value)
            if not code:
                continue
            if code not in values:
                results[missing_name].append(code)
                continue
            sheet.cell(row, month_col).value = values[code]
            results[counter_name] += 1

    return results


def update_3dp_overall(workbook, raw_wb) -> dict[str, Any]:
    sheet = workbook["3dpOverall"]
    latest_values = {code: read_3dp_overall(raw_wb, code) for code in OVERALL_3DP}
    months = {month for month, _ in latest_values.values()}
    if len(months) != 1:
        raise ValueError(f"3dp overall tables do not share one latest month: {latest_values!r}")
    latest_month = months.pop()

    row = None
    last_date_row = None
    last_date = None
    for candidate in range(2, sheet.max_row + 1):
        value = sheet.cell(candidate, 1).value
        if not isinstance(value, datetime):
            continue
        last_date_row = candidate
        last_date = value
        if value.strftime("%Y-%m") == latest_month.strftime("%Y-%m"):
            row = candidate
    if row is None:
        if not isinstance(last_date, datetime) or last_date_row is None or latest_month <= last_date:
            raise ValueError(f"Cannot append 3dp overall month {latest_month:%Y-%m}")
        row = last_date_row + 1
        for col in range(1, sheet.max_column + 1):
            source = sheet.cell(last_date_row, col)
            target = sheet.cell(row, col)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
        sheet.cell(row, 1).value = latest_month

    headers = {clean_code(sheet.cell(1, col).value): col for col in range(1, sheet.max_column + 1)}
    for code, (_, value) in latest_values.items():
        if code not in headers:
            raise ValueError(f"3dpOverall is missing column {code}")
        sheet.cell(row, headers[code]).value = value

    return {"month": latest_month.strftime("%Y-%m"), "codes_updated": sorted(latest_values)}


def download_ons_file(url: str, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / f"consumerpriceinflationdetailedreferencetables_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    print(f"Downloading ONS workbook to {target}...", flush=True)
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/126.0 Safari/537.36"
            ),
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/octet-stream,*/*"
            ),
        },
    )
    with urllib.request.urlopen(request) as response, target.open("wb") as output:
        shutil.copyfileobj(response, output)
    print(f"Downloaded ONS workbook: {target.stat().st_size:,} bytes", flush=True)
    return target


def fail_if_missing_codes(results: list[dict[str, Any]], allow_missing: bool) -> None:
    missing: list[str] = []
    for result in results:
        for field in ("missing_price_codes", "missing_weight_codes"):
            if result[field]:
                missing.append(f"{result['series']} {field}: {', '.join(result[field][:20])}")
    if missing and not allow_missing:
        joined = "\n".join(missing)
        raise ValueError(f"Missing codes found. Workbook not saved.\n{joined}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Update Weights And Prices.xlsx from the latest ONS detailed reference tables.")
    parser.add_argument("--ons", type=Path, help="Path to a downloaded ONS detailed reference tables workbook.")
    parser.add_argument("--download", action="store_true", help="Download the current ONS workbook before updating.")
    parser.add_argument("--ons-url", default=ONS_URL, help="ONS detailed reference tables URL.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Curated Weights And Prices workbook to update.")
    parser.add_argument("--output", type=Path, help="Optional output workbook. Defaults to updating --workbook in place.")
    parser.add_argument("--no-backup", action="store_true", help="Do not create a timestamped backup when updating in place.")
    parser.add_argument("--allow-missing", action="store_true", help="Save even if some workbook codes are missing in the ONS file.")
    args = parser.parse_args()

    if args.download:
        ons_path = download_ons_file(args.ons_url, DEFAULT_DOWNLOAD_DIR)
    elif args.ons:
        ons_path = args.ons
    else:
        raise ValueError("Provide --ons path or use --download")

    if not ons_path.exists():
        raise FileNotFoundError(ons_path)
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    print(f"Opening ONS workbook: {ons_path}", flush=True)
    raw_wb = load_workbook(ons_path, read_only=True, data_only=True)
    print(f"Opening curated workbook: {args.workbook}", flush=True)
    workbook = load_workbook(args.workbook)

    results = []
    for series in ("CPIH", "CPI", "RPI"):
        print(f"Updating {series} weights and prices...", flush=True)
        result = update_series(workbook, raw_wb, series)
        results.append(result)
        print(
            f"Updated {series} {result['month']}: "
            f"{result['prices_updated']} prices, {result['weights_updated']} weights",
            flush=True,
        )
    fail_if_missing_codes(results, args.allow_missing)
    print("Updating 3dpOverall...", flush=True)
    overall_result = update_3dp_overall(workbook, raw_wb)

    output = args.output or args.workbook
    if output == args.workbook and not args.no_backup:
        backup = args.workbook.with_name(f"{args.workbook.stem} backup {datetime.now():%Y%m%d_%H%M%S}{args.workbook.suffix}")
        shutil.copy2(args.workbook, backup)
        print(f"Backup written: {backup}", flush=True)

    print(f"Saving updated workbook: {output}", flush=True)
    workbook.save(output)
    print("Save complete.", flush=True)

    print(f"Updated workbook: {output}")
    for result in results:
        print(
            f"{result['series']} {result['month']}: "
            f"{result['prices_updated']} prices, {result['weights_updated']} weights"
        )
    print(f"3dpOverall {overall_result['month']}: {', '.join(overall_result['codes_updated'])}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
