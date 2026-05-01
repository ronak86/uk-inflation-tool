from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Weights And Prices.xlsx"
OUTPUT = ROOT / "web" / "data" / "cpih.json"
JS_OUTPUT = ROOT / "web" / "data" / "cpih-data.js"


def iso_month(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    raise ValueError(f"Unexpected month header: {value!r}")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def clean_code(value):
    value = clean(value)
    if value in (None, "", "#N/A"):
        return None
    return str(value).strip()


def item_prefix(name):
    if not name:
        return ""
    return str(name).strip().split(" ", 1)[0]


def read_legacy_sector_defs(workbook) -> dict:
    rows = list(workbook["SectorDefs"].iter_rows(values_only=True))
    sector_defs = {
        "CPI": {"boeCodes": set(), "servicesCodes": set(), "nonCoreCodes": set()},
        "CPIH": {"boeCodes": set(), "servicesCodes": set(), "nonCoreCodes": set()},
        "RPI": {"boeCodes": set(), "servicesCodes": set(), "nonCoreCodes": set(), "housingCodes": set()},
    }

    for row in rows[1:]:
        for code in (row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None):
            code = clean_code(code)
            if code:
                sector_defs["CPI"]["boeCodes"].add(code)
        for code in (row[3] if len(row) > 3 else None, row[4] if len(row) > 4 else None):
            code = clean_code(code)
            if code:
                sector_defs["CPIH"]["boeCodes"].add(code)

        for code in (row[7] if len(row) > 7 else None, row[8] if len(row) > 8 else None):
            code = clean_code(code)
            if code:
                sector_defs["CPI"]["servicesCodes"].add(code)
        for code in (row[9] if len(row) > 9 else None, row[10] if len(row) > 10 else None):
            code = clean_code(code)
            if code:
                sector_defs["CPIH"]["servicesCodes"].add(code)
        for code in (row[12] if len(row) > 12 else None, row[13] if len(row) > 13 else None):
            code = clean_code(code)
            if code:
                sector_defs["RPI"]["servicesCodes"].add(code)

        for code in (row[16] if len(row) > 16 else None, row[17] if len(row) > 17 else None):
            code = clean_code(code)
            if code:
                sector_defs["CPI"]["nonCoreCodes"].add(code)
        for code in (row[18] if len(row) > 18 else None, row[19] if len(row) > 19 else None):
            code = clean_code(code)
            if code:
                sector_defs["CPIH"]["nonCoreCodes"].add(code)
        for code in (row[21] if len(row) > 21 else None, row[22] if len(row) > 22 else None):
            code = clean_code(code)
            if code:
                sector_defs["RPI"]["nonCoreCodes"].add(code)

    return sector_defs


def read_definitions(workbook) -> dict:
    rows = list(workbook["Definitions"].iter_rows(values_only=True))
    sector_defs = {
        "CPIH": {"boeCodes": set(), "exBoeCodes": set(), "servicesCodes": set(), "nonCoreCodes": set(), "housingCodes": set()},
        "CPI": {"boeCodes": set(), "exBoeCodes": set(), "servicesCodes": set(), "nonCoreCodes": set(), "housingCodes": set()},
        "RPI": {"boeCodes": set(), "exBoeCodes": set(), "servicesCodes": set(), "nonCoreCodes": set(), "housingCodes": set()},
    }
    blocks = {
        "CPIH": {"weight": 1, "price": 2, "sector": 3, "core": 4, "boe": 5},
        "CPI": {"weight": 9, "price": 10, "sector": 11, "core": 12, "boe": 13},
        "RPI": {"weight": 16, "price": 17, "sector": 18, "core": 19, "boe": None},
    }

    for row in rows[1:]:
        for series, columns in blocks.items():
            codes = (
                row[columns["weight"]] if len(row) > columns["weight"] else None,
                row[columns["price"]] if len(row) > columns["price"] else None,
            )
            clean_codes = [code for code in (clean_code(value) for value in codes) if code]
            if not clean_codes:
                continue

            sector = clean(row[columns["sector"]] if len(row) > columns["sector"] else None)
            core = clean(row[columns["core"]] if len(row) > columns["core"] else None)
            boe = clean(row[columns["boe"]] if columns["boe"] is not None and len(row) > columns["boe"] else None)

            if sector == "Services":
                sector_defs[series]["servicesCodes"].update(clean_codes)
            elif sector == "Housing":
                sector_defs[series]["housingCodes"].update(clean_codes)

            if core == "Non Core":
                sector_defs[series]["nonCoreCodes"].update(clean_codes)

            if boe == "BoE Custom Services":
                sector_defs[series]["boeCodes"].update(clean_codes)
            elif boe == "Ex BoE Custom Services":
                sector_defs[series]["exBoeCodes"].update(clean_codes)

    return sector_defs


def read_sector_defs(workbook) -> dict:
    if "Definitions" in workbook.sheetnames:
        return read_definitions(workbook)
    if "SectorDefs" in workbook.sheetnames:
        return read_legacy_sector_defs(workbook)
    return {}


def annotate_sectors(payload: dict, sector_defs: dict) -> None:
    series_defs = sector_defs.get(payload["series"], {})
    boe_codes = series_defs.get("boeCodes", set())
    ex_boe_codes = series_defs.get("exBoeCodes", set())
    services_codes = series_defs.get("servicesCodes", set())
    non_core_codes = series_defs.get("nonCoreCodes", set())
    housing_codes = series_defs.get("housingCodes", set())

    for item in payload["items"]:
        codes = {clean_code(item["weightCode"]), clean_code(item["priceCode"])}
        codes.discard(None)
        item["sectors"] = {
            "boe": bool(codes & boe_codes),
            "exBoe": bool(codes & ex_boe_codes),
            "services": bool(codes & services_codes),
            "nonCore": bool(codes & non_core_codes),
            "housing": bool(codes & housing_codes),
        }


def read_series(workbook, series: str) -> dict:
    weights = workbook[f"{series}_Weights"]
    prices = workbook[f"{series}_Prices"]

    weight_rows = list(weights.iter_rows(values_only=True))
    price_rows = list(prices.iter_rows(values_only=True))

    if len(weight_rows) != len(price_rows):
        raise ValueError("Weight and price sheets have different row counts")

    months = [iso_month(value) for value in weight_rows[0][4:]]
    items = []

    for weight_row, price_row in zip(weight_rows[1:], price_rows[1:]):
        weight_key = tuple(clean(value) for value in weight_row[:4])
        price_key = tuple(clean(value) for value in price_row[:4])
        if weight_key != price_key:
            raise ValueError(f"Row mismatch: {weight_key!r} != {price_key!r}")

        name, level, weight_code, price_code = weight_key
        items.append(
            {
                "name": name,
                "level": int(level),
                "weightCode": weight_code,
                "priceCode": price_code,
                "weights": [float(value) for value in weight_row[4:]],
                "prices": [float(value) for value in price_row[4:]],
            }
        )

    return {
        "series": series,
        "sourceWorkbook": SOURCE.name,
        "months": months,
        "items": items,
    }


def read_overall_3dp(workbook) -> dict:
    sheet = workbook["3dpOverall"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [clean(value) for value in rows[0]]
    code_map = {"D7BT": "CPI", "L522": "CPIH"}
    code_columns = {
        series: header.index(code)
        for code, series in code_map.items()
        if code in header
    }
    if not code_columns:
        return {}

    months = []
    values = {series: [] for series in code_columns}
    for row in rows[1:]:
        if row[0] is None:
            continue
        months.append(iso_month(row[0]))
        for series, column in code_columns.items():
            value = row[column]
            values[series].append(float(value) if value is not None else None)

    return {
        series: {
            "priceCode": "D7BT" if series == "CPI" else "L522",
            "months": months,
            "prices": values[series],
        }
        for series in values
    }


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    overall_3dp = read_overall_3dp(workbook)
    sector_defs = read_sector_defs(workbook)
    series_payloads = {
        series: read_series(workbook, series)
        for series in ("CPIH", "CPI", "RPI")
    }
    for series, overall in overall_3dp.items():
        if series in series_payloads:
            series_payloads[series]["overall3dp"] = overall
    for payload in series_payloads.values():
        annotate_sectors(payload, sector_defs)

    payload = {
        "sourceWorkbook": SOURCE.name,
        "series": series_payloads,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload_json = json.dumps(payload, separators=(",", ":"))
    OUTPUT.write_text(payload_json, encoding="utf-8")
    JS_OUTPUT.write_text(f"window.CPIH_DATA={payload_json};\n", encoding="utf-8")
    summary = ", ".join(
        f"{series}: {len(data['items'])} items, {len(data['months'])} months"
        for series, data in series_payloads.items()
    )
    print(f"Wrote {OUTPUT} with {summary}")


if __name__ == "__main__":
    main()
